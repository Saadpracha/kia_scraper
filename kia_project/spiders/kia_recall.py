import scrapy
import json
import csv
import io
from html import unescape
from html.parser import HTMLParser
from pathlib import Path


class _HTMLTextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self._parts = []
        self._skip_depth = 0

    def handle_starttag(self, tag, attrs):
        if tag.lower() in ("script", "style"):
            self._skip_depth += 1

    def handle_endtag(self, tag):
        if tag.lower() in ("script", "style") and self._skip_depth:
            self._skip_depth -= 1

    def handle_data(self, data):
        if self._skip_depth == 0 and data:
            self._parts.append(data)

    def get_text(self):
        return " ".join("".join(self._parts).split())


class KiaRecallsSpider(scrapy.Spider):
    name = "kia_recalls"

    custom_settings = {
        "FEED_EXPORT_FIELDS": [
            "account",
            "vin",
            "year",
            "make",
            "model",
            "type",
            "campaign",
            "status",
            "source_url",
            "language",
            "title",
            "description",
            "other_fields",
        ],
    }

    def __init__(self, lang="en", input_file="input.csv", encoding=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.lang = lang.lower()

        if self.lang not in ["en", "fr"]:
            raise ValueError("lang must be either 'en' or 'fr'")

        self.input_file = input_file
        self.encoding = encoding

        self.base_urls = {
            "en": "https://www.kia.ca/content/marketing/ca/en/owners/recalls/jcr:content/root/container/container/sidebar_container/right/par-sidebar-container/recalls.recall.json?_vin={vin}",
            "fr": "https://www.kia.ca/content/marketing/ca/fr/owners/recalls/jcr:content/root/container/container/sidebar_container/right/par-sidebar-container/recalls.recall.json?_vin={vin}",
        }

    def _read_csv_text(self, path: Path) -> str:
        raw = path.read_bytes()
        if self.encoding:
            return raw.decode(self.encoding)

        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue

        raise UnicodeDecodeError(
            "utf-8",
            raw,
            0,
            len(raw),
            f"Could not decode {path} (tried utf-8-sig, utf-8, cp1252, latin-1)",
        )

    def _iter_xlsx_rows(self, path: Path):
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        headers = [
            str(cell).strip() if cell is not None else ""
            for cell in next(row_iter, ())
        ]

        for cells in row_iter:
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = cells[index] if index < len(cells) else None
                row[header] = str(value).strip() if value is not None else ""
            yield row

        workbook.close()

    def _iter_input_rows(self, path: Path):
        suffix = path.suffix.lower()

        if suffix in (".xlsx", ".xlsm"):
            yield from self._iter_xlsx_rows(path)
            return

        if suffix == ".csv":
            reader = csv.DictReader(io.StringIO(self._read_csv_text(path)))
            yield from reader
            return

        raise ValueError(
            f"Unsupported input file type '{suffix}'. Use .csv, .xlsx, or .xlsm."
        )

    def _get_column(self, row, column_name):
        target = column_name.strip().upper()
        for key, value in row.items():
            if key.strip().upper() == target and str(value).strip():
                return str(value).strip()
        return ""

    def _get_vin(self, row):
        return self._get_column(row, "VIN")

    def _get_account(self, row):
        return self._get_column(row, "ACCOUNT")

    def start_requests(self):
        input_path = Path(self.input_file)

        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_file}")

        account = ""
        for row in self._iter_input_rows(input_path):
            if not account:
                account = self._get_account(row)

            vin = self._get_vin(row)

            if not vin:
                continue

            url = self.base_urls[self.lang].format(vin=vin)

            yield scrapy.Request(
                url=url,
                callback=self.parse,
                meta={
                    "vin": vin,
                    "account": account,
                },
            )

    def clean_html(self, text):
        if not text:
            return ""

        parser = _HTMLTextExtractor()
        parser.feed(unescape(text))
        parser.close()
        return parser.get_text()

    def map_status(self, status):
        status = status.strip().lower()

        if self.lang == "en":
            mapping = {
                "open": "Outstanding",
                "closed": "Completed",
            }
        else:
            mapping = {
                "open": "À effectuer",
                "closed": "Réaliser",
            }

        return mapping.get(status, status)

    def _build_other_fields(self, extra):
        if not extra:
            return "{}"
        return json.dumps(extra, ensure_ascii=False)

    def _recall_extra_fields(self, recall):
        used_keys = {
            "modelYear",
            "model",
            "recallId",
            "status",
            "recallTitleEn",
            "recallTitleFr",
            "descriptionEN",
            "descriptionFR",
        }
        return {
            key: value
            for key, value in recall.items()
            if key not in used_keys and value not in (None, "")
        }

    def _make_item(
        self,
        vin,
        source_url,
        *,
        account="",
        year="",
        model="",
        campaign="",
        status="",
        title="",
        description="",
        other_fields=None,
    ):
        return {
            "account": account,
            "vin": vin,
            "year": year,
            "make": "kia",
            "model": model,
            "type": "safety_recall",
            "campaign": campaign,
            "status": status,
            "source_url": source_url,
            "language": self.lang,
            "title": title,
            "description": description,
            "other_fields": other_fields if other_fields is not None else "{}",
        }

    def parse(self, response):
        vin = response.meta["vin"]
        account = response.meta.get("account", "")
        source_url = response.url

        data = json.loads(response.text)

        if data.get("isVinValid") != "true":
            yield self._make_item(
                vin,
                source_url,
                account=account,
                title="Invalid VIN",
                other_fields=self._build_other_fields(
                    {"isVinValid": data.get("isVinValid", "")}
                ),
            )
            return

        recalls = data.get("recallInfo", [])

        for recall in recalls:

            if self.lang == "en":
                title = recall.get("recallTitleEn", "")
                description = recall.get("descriptionEN", "")
            else:
                title = recall.get("recallTitleFr", "")
                description = recall.get("descriptionFR", "")

            yield self._make_item(
                vin,
                source_url,
                account=account,
                year=recall.get("modelYear", ""),
                model=recall.get("model", ""),
                campaign=recall.get("recallId", ""),
                title=self.clean_html(title),
                description=self.clean_html(description),
                status=self.map_status(recall.get("status", "")),
                other_fields=self._build_other_fields(
                    self._recall_extra_fields(recall)
                ),
            )